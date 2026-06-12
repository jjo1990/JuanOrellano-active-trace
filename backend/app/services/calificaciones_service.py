import csv
import io
import logging
import uuid
from collections.abc import Sequence
from datetime import datetime, timedelta, timezone

from fastapi import HTTPException, UploadFile
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import action_codes
from app.models.calificacion import Calificacion
from app.models.entrada_padron import EntradaPadron
from app.models.version_padron import VersionPadron
from app.models.umbral_materia import UmbralMateria
from app.repositories.calificaciones_repository import CalificacionesRepository
from app.schemas.calificaciones import (
    ColumnaDetectada,
    ConfirmCalificacionesResponse,
    ImportCalificacionesPreviewResponse,
    ImportFinalizacionResponse,
    UmbralMateriaDTO,
    UmbralMateriaUpdate,
)

logger = logging.getLogger(__name__)


class CalificacionesService:
    _preview_cache: dict[str, dict] = {}
    _PREVIEW_TTL = timedelta(minutes=30)

    def __init__(self, session: AsyncSession, tenant_id: uuid.UUID) -> None:
        self._repo = CalificacionesRepository(session, tenant_id)
        self._session = session
        self._tenant_id = tenant_id

    async def _run(self, coro):
        try:
            return await coro
        except IntegrityError as exc:
            await self._session.rollback()
            logger.warning("IntegrityError: %s", exc)
            raise HTTPException(status_code=409, detail="Conflicto de integridad.") from exc

    async def _commit(self) -> None:
        try:
            await self._session.commit()
        except IntegrityError as exc:
            await self._session.rollback()
            logger.warning("IntegrityError: %s", exc)
            raise HTTPException(status_code=409, detail="Conflicto de integridad.") from exc

    # ── Parseo de archivos ────────────────────────────────────────────────

    @staticmethod
    def _detect_format(filename: str) -> str:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xlsx":
            return "xlsx"
        if ext == "csv":
            return "csv"
        raise HTTPException(status_code=422, detail=f"Formato no soportado: .{ext}")

    async def _parse_csv(self, file: UploadFile) -> list[dict]:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=422, detail="El archivo CSV está vacío.")
        return rows

    async def _parse_xlsx(self, file: UploadFile) -> list[dict]:
        from openpyxl import load_workbook
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=422, detail="El archivo XLSX no tiene hojas.")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            raise HTTPException(status_code=422, detail="El archivo XLSX está vacío.")
        rows = []
        for row in rows_iter:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ""
            if any(row_dict.values()):
                rows.append(row_dict)
        wb.close()
        return rows

    # ── Detección de columnas ──────────────────────────────────────────────

    def _detect_column_types(self, headers: list[str]) -> list[ColumnaDetectada]:
        columns = []
        for h in headers:
            if h.endswith("(Real)"):
                columns.append(ColumnaDetectada(
                    name=h, tipo="numerica", aprobatorio=False, actividad=h,
                ))
            elif h in ("Satisfactorio", "Supera lo esperado"):
                columns.append(ColumnaDetectada(
                    name=h, tipo="textual", aprobatorio=True, actividad=h,
                ))
            else:
                columns.append(ColumnaDetectada(
                    name=h, tipo="textual", aprobatorio=False, actividad=h,
                ))
        return columns

    # ── Derivación de aprobado ─────────────────────────────────────────────

    @staticmethod
    def _es_aprobado(cal: Calificacion, umbral: UmbralMateria | None) -> bool:
        if umbral is not None:
            pct = umbral.umbral_pct
            valores = umbral.valores_aprobatorios
        else:
            pct = 60
            valores = ["Satisfactorio", "Supera lo esperado"]

        if cal.nota_numerica is not None:
            return cal.nota_numerica >= pct
        if cal.nota_textual is not None:
            return cal.nota_textual in valores
        return False

    # ── Preview Cache ─────────────────────────────────────────────────────

    @classmethod
    def _cleanup_cache(cls) -> None:
        now = datetime.now(timezone.utc)
        expired = [
            k for k, v in cls._preview_cache.items()
            if v.get("_created_at", now) < now - cls._PREVIEW_TTL
        ]
        for k in expired:
            cls._preview_cache.pop(k, None)

    # ── Matching ──────────────────────────────────────────────────────────

    async def _match_entradas(self, rows: list[dict]) -> dict:
        """Match rows against EntradaPadron in active versions by nombre+apellidos."""
        active_students = await self._get_active_entradas()
        name_index = {}
        for e in active_students:
            key = f"{e.nombre.lower().strip()}|{e.apellidos.lower().strip()}"
            name_index[key] = e

        matched = []
        ignored = []
        for row in rows:
            nombre = row.get("Nombre", row.get("nombre", "")).strip()
            apellidos = row.get("Apellidos", row.get("apellidos", "")).strip()
            key = f"{nombre.lower()}|{apellidos.lower()}"
            entrada = name_index.get(key)
            if entrada:
                matched.append((row, entrada))
            else:
                ignored.append(row)

        return {"matched": matched, "ignored": ignored}

    async def _get_active_entradas(self) -> Sequence[EntradaPadron]:
        from sqlalchemy import select
        stmt = (
            select(EntradaPadron)
            .join(VersionPadron, VersionPadron.id == EntradaPadron.version_id)
            .where(
                VersionPadron.activa.is_(True),
                EntradaPadron.tenant_id == self._tenant_id,
                EntradaPadron.deleted_at.is_(None),
                VersionPadron.deleted_at.is_(None),
            )
        )
        result = await self._session.execute(stmt)
        return result.scalars().all()

    # ── Import pipeline ───────────────────────────────────────────────────

    async def import_preview(
        self,
        file: UploadFile,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        asignacion_id: uuid.UUID,
        usuario_id: uuid.UUID,
    ) -> ImportCalificacionesPreviewResponse:
        self._cleanup_cache()
        fmt = self._detect_format(file.filename or "archivo.csv")
        if fmt == "xlsx":
            rows = await self._parse_xlsx(file)
        else:
            rows = await self._parse_csv(file)
        if not rows:
            raise HTTPException(status_code=422, detail="El archivo no contiene datos.")
        headers = list(rows[0].keys())
        columns = self._detect_column_types(headers)
        actividades = [
            {"name": c.name, "tipo": c.tipo, "aprobatorio": c.aprobatorio}
            for c in columns
            if c.name not in ("Nombre", "Apellidos", "nombre", "apellidos")
        ]
        preview_token = str(uuid.uuid4())
        self._preview_cache[preview_token] = {
            "filename": file.filename or "archivo",
            "rows": rows,
            "materia_id": str(materia_id),
            "cohorte_id": str(cohorte_id),
            "asignacion_id": str(asignacion_id),
            "usuario_id": str(usuario_id),
            "_created_at": datetime.now(timezone.utc),
            "columns": [c.model_dump() for c in columns],
        }
        preview_rows = []
        for row in rows[:5]:
            preview_row = {k: str(v) for k, v in row.items()}
            preview_rows.append(preview_row)
        return ImportCalificacionesPreviewResponse(
            preview_token=preview_token,
            filename=file.filename or "archivo",
            detected_rows=len(rows),
            columns=columns,
            actividades=actividades,
            preview_rows=preview_rows,
        )

    async def confirm_import(
        self,
        preview_token: str,
        actividades_seleccionadas: list[str],
    ) -> ConfirmCalificacionesResponse:
        preview = self._preview_cache.pop(preview_token, None)
        if preview is None:
            raise HTTPException(status_code=404, detail="Token de preview no válido o expirado.")
        rows = preview["rows"]
        materia_id = uuid.UUID(preview["materia_id"])
        cohorte_id = uuid.UUID(preview["cohorte_id"])
        usuario_id = uuid.UUID(preview["usuario_id"])

        match_result = await self._match_entradas(rows)
        matched_pairs = match_result["matched"]
        ignored_count = len(match_result["ignored"])

        calificaciones = []
        for row, entrada in matched_pairs:
            for actividad in actividades_seleccionadas:
                valor = row.get(actividad, "")
                nota_numerica = None
                nota_textual = None
                col_type = None
                for col in preview["columns"]:
                    if col["name"] == actividad:
                        col_type = col["tipo"]
                        break
                if col_type == "numerica":
                    try:
                        nota_numerica = float(valor) if valor else None
                    except ValueError:
                        continue
                else:
                    nota_textual = valor if valor else None

                cal = Calificacion(
                    tenant_id=self._tenant_id,
                    entrada_padron_id=entrada.id,
                    materia_id=materia_id,
                    actividad=actividad,
                    nota_numerica=nota_numerica,
                    nota_textual=nota_textual,
                    origen="Importado",
                )
                calificaciones.append(cal)

        if calificaciones:
            await self._repo.create_calificaciones_batch(calificaciones)

        _log_audit_calificaciones(
            self._session, usuario_id, self._tenant_id,
            materia_id, len(calificaciones),
        )
        await self._commit()

        return ConfirmCalificacionesResponse(
            calificaciones_count=len(calificaciones),
            ignorados_count=ignored_count,
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            importado_at=datetime.now(timezone.utc),
        )

    async def import_finalizacion(
        self,
        file: UploadFile,
        materia_id: uuid.UUID,
        cohorte_id: uuid.UUID,
        asignacion_id: uuid.UUID,
        usuario_id: uuid.UUID,
    ) -> ImportFinalizacionResponse:
        fmt = self._detect_format(file.filename or "archivo.csv")
        if fmt == "xlsx":
            rows = await self._parse_xlsx(file)
        else:
            rows = await self._parse_csv(file)
        if not rows:
            raise HTTPException(status_code=422, detail="El archivo no contiene datos.")

        match_result = await self._match_entradas(rows)
        matched_pairs = match_result["matched"]
        ignored = match_result["ignored"]

        calificaciones = []
        for row, entrada in matched_pairs:
            estado = row.get("Estado", row.get("estado", ""))
            if estado in ("Finalizado", "Completo"):
                cal = Calificacion(
                    tenant_id=self._tenant_id,
                    entrada_padron_id=entrada.id,
                    materia_id=materia_id,
                    actividad="Finalización",
                    nota_numerica=None,
                    nota_textual=estado,
                    origen="Importado",
                )
                calificaciones.append(cal)
            else:
                ignored.append(row)

        if calificaciones:
            await self._repo.create_calificaciones_batch(calificaciones)

        _log_audit_calificaciones(
            self._session, usuario_id, self._tenant_id,
            materia_id, len(calificaciones),
        )
        await self._commit()

        return ImportFinalizacionResponse(
            calificaciones_count=len(calificaciones),
            ignorados_count=len(ignored),
            materia_id=materia_id,
        )

    # ── Umbral CRUD ──────────────────────────────────────────────────────

    async def get_umbral(
        self, materia_id: uuid.UUID, asignacion_id: uuid.UUID,
    ) -> UmbralMateriaDTO:
        umbral = await self._repo.get_umbral(materia_id, asignacion_id)
        if umbral is not None:
            return UmbralMateriaDTO(
                id=umbral.id,
                asignacion_id=umbral.asignacion_id,
                materia_id=umbral.materia_id,
                umbral_pct=umbral.umbral_pct,
                valores_aprobatorios=umbral.valores_aprobatorios,
            )
        return UmbralMateriaDTO(
            id=uuid.uuid4(),
            asignacion_id=asignacion_id,
            materia_id=materia_id,
            umbral_pct=60,
            valores_aprobatorios=["Satisfactorio", "Supera lo esperado"],
        )

    async def upsert_umbral(
        self, materia_id: uuid.UUID, asignacion_id: uuid.UUID, data: UmbralMateriaUpdate,
    ) -> UmbralMateriaDTO:
        umbral = UmbralMateria(
            tenant_id=self._tenant_id,
            asignacion_id=asignacion_id,
            materia_id=materia_id,
            umbral_pct=data.umbral_pct,
            valores_aprobatorios=data.valores_aprobatorios,
        )
        result = await self._run(self._repo.upsert_umbral(umbral))
        await self._commit()
        return UmbralMateriaDTO(
            id=result.id,
            asignacion_id=result.asignacion_id,
            materia_id=result.materia_id,
            umbral_pct=result.umbral_pct,
            valores_aprobatorios=result.valores_aprobatorios,
        )


def _log_audit_calificaciones(
    session: AsyncSession,
    usuario_id: uuid.UUID,
    tenant_id: uuid.UUID,
    materia_id: uuid.UUID,
    filas: int,
) -> None:
    from app.models.audit_log import AuditLog
    record = AuditLog(
        actor_id=usuario_id,
        tenant_id=tenant_id,
        accion=action_codes.CALIFICACIONES_IMPORTAR,
        detalle={"materia_id": str(materia_id)},
        filas_afectadas=filas,
        materia_id=materia_id,
    )
    session.add(record)
