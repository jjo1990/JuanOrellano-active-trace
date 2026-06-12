import csv
import io
import logging
import uuid
from collections.abc import Sequence
from datetime import datetime, timedelta, timezone

from fastapi import HTTPException, UploadFile
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import action_codes
from app.models.entrada_padron import EntradaPadron
from app.models.version_padron import VersionPadron
from app.repositories.padron_repository import PadronRepository
from app.schemas.padron import (
    ConfirmResponse,
    EntradaPadronDTO,
    ImportPreviewResponse,
    VaciarResponse,
    VersionPadronDTO,
)

logger = logging.getLogger(__name__)


class PadronService:
    """Service de importación y gestión de padrones."""

    _preview_cache: dict[str, dict] = {}
    _PREVIEW_TTL = timedelta(minutes=30)

    def __init__(self, session: AsyncSession, tenant_id: uuid.UUID) -> None:
        self._repo = PadronRepository(session, tenant_id)
        self._session = session
        self._tenant_id = tenant_id

    async def _run(self, coro):
        try:
            return await coro
        except IntegrityError as exc:
            await self._session.rollback()
            logger.warning("IntegrityError: %s", exc)
            raise HTTPException(status_code=409, detail="Conflicto de integridad.") from exc

    async def _commit(self) -> None:
        try:
            await self._session.commit()
        except IntegrityError as exc:
            await self._session.rollback()
            logger.warning("IntegrityError: %s", exc)
            raise HTTPException(status_code=409, detail="Conflicto de integridad.") from exc

    # ── Parseo de archivos ────────────────────────────────────────────────

    @staticmethod
    def _detect_format(filename: str) -> str:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xlsx":
            return "xlsx"
        if ext == "csv":
            return "csv"
        raise HTTPException(status_code=422, detail=f"Formato no soportado: .{ext}")

    @staticmethod
    def _validate_columns(headers: set[str], required: set[str]) -> None:
        missing = required - headers
        if missing:
            raise HTTPException(
                status_code=422,
                detail=f"Columnas requeridas faltantes: {', '.join(sorted(missing))}",
            )

    async def _parse_csv(self, file: UploadFile) -> list[dict]:
        content = await file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=422, detail="El archivo CSV está vacío.")
        return rows

    async def _parse_xlsx(self, file: UploadFile) -> list[dict]:
        from openpyxl import load_workbook
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=422, detail="El archivo XLSX no tiene hojas.")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            raise HTTPException(status_code=422, detail="El archivo XLSX está vacío.")
        rows = []
        for row in rows_iter:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ""
            if any(row_dict.values()):
                rows.append(row_dict)
        wb.close()
        return rows

    # ── Preview Cache ─────────────────────────────────────────────────────

    @classmethod
    def _cleanup_cache(cls) -> None:
        now = datetime.now(timezone.utc)
        expired = [k for k, v in cls._preview_cache.items() if v.get("_created_at", now) < now - cls._PREVIEW_TTL]
        for k in expired:
            cls._preview_cache.pop(k, None)

    # ── Import pipeline ───────────────────────────────────────────────────

    async def import_preview(
        self, file: UploadFile, materia_id: uuid.UUID, cohorte_id: uuid.UUID, usuario_id: uuid.UUID,
    ) -> ImportPreviewResponse:
        self._cleanup_cache()
        fmt = self._detect_format(file.filename or "archivo.csv")
        required = {"nombre", "apellidos"}
        if fmt == "xlsx":
            rows = await self._parse_xlsx(file)
        else:
            rows = await self._parse_csv(file)
        if not rows:
            raise HTTPException(status_code=422, detail="El archivo no contiene datos.")
        headers_set = set(rows[0].keys())
        self._validate_columns(headers_set, required)
        preview_token = str(uuid.uuid4())
        self._preview_cache[preview_token] = {
            "filename": file.filename or "archivo",
            "rows": rows,
            "materia_id": str(materia_id),
            "cohorte_id": str(cohorte_id),
            "usuario_id": str(usuario_id),
            "_created_at": datetime.now(timezone.utc),
            "columns": list(rows[0].keys()),
        }
        return ImportPreviewResponse(
            preview_token=preview_token,
            filename=file.filename or "archivo",
            detected_rows=len(rows),
            columns=list(rows[0].keys()),
            preview_rows=rows[:5],
        )

    async def confirm_import(self, preview_token: str) -> ConfirmResponse:
        preview = self._preview_cache.pop(preview_token, None)
        if preview is None:
            raise HTTPException(status_code=404, detail="Token de preview no válido o expirado.")
        rows = preview["rows"]
        materia_id = uuid.UUID(preview["materia_id"])
        cohorte_id = uuid.UUID(preview["cohorte_id"])
        usuario_id = uuid.UUID(preview["usuario_id"])

        version = VersionPadron(
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            cargado_por=usuario_id,
            activa=True,
        )
        result = await self._run(self._repo.create_version(version))
        entradas = []
        for row in rows:
            entrada = EntradaPadron(
                version_id=result.id,
                nombre=row.get("nombre", ""),
                apellidos=row.get("apellidos", ""),
                email=row.get("email"),
                comision=row.get("comision"),
                regional=row.get("regional"),
            )
            entradas.append(entrada)
        await self._repo.create_entradas_batch(entradas)
        _log_audit(self._session, usuario_id, self._tenant_id, result.id, materia_id, len(entradas))
        await self._commit()
        await self._session.refresh(result)

        return ConfirmResponse(
            version_id=result.id,
            entries_count=len(entradas),
            materia_id=materia_id,
            cohorte_id=cohorte_id,
            cargado_at=datetime.now(timezone.utc),
        )

    async def vaciar_materia(self, materia_id: uuid.UUID, usuario_id: uuid.UUID) -> VaciarResponse:
        filas = await self._run(self._repo.soft_delete_by_materia(materia_id))
        _log_audit(self._session, usuario_id, self._tenant_id, None, materia_id, filas)
        await self._commit()
        return VaciarResponse(
            materia_id=materia_id,
            filas_afectadas=filas,
            mensaje=f"Se eliminaron {filas} registros de la materia.",
        )

    async def get_entradas(self, version_id: uuid.UUID) -> list[EntradaPadronDTO]:
        entradas = await self._repo.list_entradas(version_id)
        return [
            EntradaPadronDTO(
                id=e.id,
                version_id=e.version_id,
                usuario_id=e.usuario_id,
                nombre=e.nombre,
                apellidos=e.apellidos,
                email=e.email,
                comision=e.comision,
                regional=e.regional,
            )
            for e in entradas
        ]

    async def list_versions(self, materia_id: uuid.UUID | None = None) -> list[VersionPadronDTO]:
        versions = await self._repo.list_versions(materia_id=materia_id)
        return [
            VersionPadronDTO(
                id=v.id, materia_id=v.materia_id, cohorte_id=v.cohorte_id,
                cargado_por=v.cargado_por, cargado_at=v.cargado_at, activa=v.activa,
            )
            for v in versions
        ]


def _log_audit(
    session: AsyncSession, usuario_id: uuid.UUID, tenant_id: uuid.UUID,
    version_id: uuid.UUID | None, materia_id: uuid.UUID, filas: int,
) -> None:
    from app.models.audit_log import AuditLog
    record = AuditLog(
        actor_id=usuario_id,
        tenant_id=tenant_id,
        accion=action_codes.PADRON_CARGAR,
        detalle={"version_id": str(version_id)} if version_id else None,
        filas_afectadas=filas,
        materia_id=materia_id,
    )
    session.add(record)
